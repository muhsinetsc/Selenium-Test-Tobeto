from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
from selenium.webdriver.common.action_chains import ActionChains 
import pytest
import json
from openpyxl import load_workbook
from constants.globalConstants import *
from selenium.webdriver.remote.webelement import WebElement 

def verileriKullanJson(dosya_yolu):
        with open(dosya_yolu, "r", encoding='utf-8') as dosya:
            data = json.load(dosya)
        return data 

class Test_Demo:
    #pytest tarafından tanımlanan bir method 
    #her test öncesi otomatik olarak çalıştırılır
    def precondition(self):
        driver =webdriver.Chrome()
        driver.maximize_window()
        driver.get(BASE_URL)  
        return driver
    
    def teardown_method(self):
        driver = self.precondition()
        driver.quit()

    @pytest.mark.skip #tüm testler koşulurken "skip" şeklinde işaretlenen testlerimi atla
    def test_demo(self):
        print("x")
        text = "Hello"
        assert text == "Hello"

    # def readInvalidDataFromExcel():
    #     return [("1","1"),("locked_out_user",secret_sauce),(standard_user,secret_sauce)]
        
    @staticmethod
    def readInvalidDataFromExcel():
        excelFile = load_workbook("C:/Users/hasine_tasci/OneDrive/Masaüstü/Selenium-Test-Tobeto/PyTest/data/invalidLogin.xlsx")
        # sheet = excelFile.active
        sheet = excelFile["Sheet1"] #hangi sayfasada çalıştığımızı gösteriyor.
        rows = sheet.max_row #kaçıncı satıra kadar benim verim var
        data = [] #dataiàsminde boş bir dizi oluşturduk.
        for i in range(2,rows+1):
            username = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((username,password))
        return data
       
    def waitForElementVisible(self,driver,locator,timeout=3):
        element = WebDriverWait(driver,timeout).until(ec.visibility_of_element_located(locator))
        return element
            
    def test_null_value(self):
        driver = self.precondition()
        loginButton = self.waitForElementVisible(driver,((By.ID,login_button_id)))
        loginButton.click()
        expectedMessage1 = self.waitForElementVisible(driver,((By.XPATH,errorMessage_xpath1)))
        testResult = expectedMessage1.text == expectedMessage1_text
        print(f"{expectedMessage1_text} {mesaj} = {testResult}")

    @pytest.mark.parametrize("excel", readInvalidDataFromExcel())
    def test_null_password(self, excel):
        username,password = excel 
        driver = self.precondition()
        userNameInput = self.waitForElementVisible(driver,(By.ID,username_id))
        userNameInput.send_keys(username)
        loginButton = self.waitForElementVisible(driver,(By.ID,login_button_id))
        loginButton.click()
        expectedMessage2 =driver.find_element(By.XPATH,expectedMessage2_xpath)
        testResult = expectedMessage2.text == expectedMessage2_text
        print(f"{expectedMessage2_text} {mesaj} = {testResult}")

    @pytest.mark.parametrize("username,password",readInvalidDataFromExcel())
    def test_user_locked(self,username,password):
        if username != locket_out or password != secret_sauce:
            pytest.skip()
        driver = self.precondition()
        userNameInput = self.waitForElementVisible(driver,(By.ID,username_id))
        passwordInput = self.waitForElementVisible(driver,(By.ID,password_id))
        actions = ActionChains(driver)
        actions.send_keys_to_element(userNameInput,username)
        actions.send_keys_to_element(userNameInput,username)
        userNameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginButton = self.waitForElementVisible(driver,(By.ID,login_button_id))
        loginButton.click()
        expectedMessage3 = self.waitForElementVisible(driver,(By.XPATH,errorMessage_xpath1))
        assert expectedMessage3.text == expectedMessage3_text
        # testResult = expectedMessage3.text == expectedMessage3_text
        # print(f"{expectedMessage3_text} {mesaj} = {testResult}")


    @pytest.mark.parametrize("username,password",readInvalidDataFromExcel())
    def test_invalid_login(self,username,password):
        if username != bir and password != bir:
            pytest.skip()
        driver = self.precondition()
        userNameInput = self.waitForElementVisible(driver,(By.ID,username_id))
        passwordInput = self.waitForElementVisible(driver,(By.ID,password_id))
        userNameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginButton = self.waitForElementVisible(driver,(By.ID,login_button_id))
        loginButton.click()
        errorMessage =self.waitForElementVisible(driver,(By.XPATH,errorMessage_xpath1))
        assert errorMessage.text == errorMessage_text, f"Beklenen hata mesajı bulunamadı. Görüntülenen mesaj: {errorMessage.text}"
        # testResult = errorMessage.text == errorMessage_text
        print(f"{errorMessage_text} {mesaj}")
        # assert errorMessage.text == errorMessage_text

    # @pytest.mark.parametrize("username,","password",verileriKullanJson())
    # @pytest.mark.parametrize("username,password", verileriKullanJson()
    # ["test_valid_login"].items())
    # @pytest.mark.parametrize("username","password",verileriKullanJson,
    # verileriKullanJson("C:/Users/Admin/Desktop/tobetoSeleniumTest/PyTestOdev/data/veri.json")['test_valid_login'].items())
    # @pytest.mark.parametrize("username, password", [(veri['username'], veri['password']) 
    # for veri in verileriKullanJson("C:/Users/Admin/Desktop/tobetoSeleniumTest/PyTestOdev/data/veri.json").values()])
    
    def test_valid_login(self):
        data = verileriKullanJson("C:/Users/hasine_tasci/OneDrive/Masaüstü/Selenium-Test-Tobeto/PyTest/data/veri.json")
        test_data = data.get("test_valid_login", {})  # JSON dosyasından verileri çeker
        username = test_data.get("username", "")  # Kullanıcı adını al
        password = test_data.get("password", "")  # Şifreyi al
        driver = self.precondition()
        userNameInput = self.waitForElementVisible(driver,(By.ID,username_id))
        passwordInput =self.waitForElementVisible(driver,(By.ID,password_id))
        actions = ActionChains(driver)
        actions.send_keys_to_element(userNameInput,username)
        actions.send_keys_to_element(passwordInput,password)
        actions.perform() #depoladığım aksiyonları çalıştır
        loginButton = self.waitForElementVisible(driver,(By.ID,login_button_id))
        loginButton.click()
        baslik1 = self.waitForElementVisible(driver,(By.XPATH,baslik_xpath))
        assert baslik1.text == baslik1_text

    def test_products_list3(self):
        data = verileriKullanJson("C:/Users/hasine_tasci/OneDrive/Masaüstü/Selenium-Test-Tobeto/PyTest/data/veri.json")
        test_data = data.get("test_valid_login", {})  # JSON dosyasından verileri çeker
        username = test_data.get("username", "")  # Kullanıcı adını al
        password = test_data.get("password", "")  # Şifreyi al
        driver = self.precondition()
        userNameInput = self.waitForElementVisible(driver,(By.ID,username_id))
        passwordInput = self.waitForElementVisible(driver,(By.ID,password_id))
        actions = ActionChains(driver)
        actions.send_keys_to_element(userNameInput,username)
        actions.send_keys_to_element(passwordInput,password)
        actions.perform() #depoladığım aksiyonları çalıştır
        loginButton = self.waitForElementVisible(driver,(By.ID,login_button_id))
        loginButton.click()
        urunEkle1 = self.waitForElementVisible(driver,(By.XPATH,urunEkle1_xpath ))
        urunEkle2 = self.waitForElementVisible(driver,(By.XPATH,urunEkle2_xpath))
        driver.execute_script("window.scrollTo(0,500)")
        actions.click(urunEkle1)
        actions.click(urunEkle2)
        sleep(2)
        actions.perform()
        sleep(2)
        sepet = self.waitForElementVisible(driver,(By.CLASS_NAME,sepet_class_name))
        actions.click(sepet).perform()
        actions.perform()
        removeButton = self.waitForElementVisible(driver,(By.XPATH,remove_button_xpath))
        actions.click(removeButton)
        checkout = self.waitForElementVisible(driver,(By.ID,checkout_id))
        actions.click(checkout)
        actions.perform()
        firstName = self.waitForElementVisible(driver,(By.ID,first_name_id))
        lastName = self.waitForElementVisible(driver,(By.ID,last_name_id))
        zipCode = self.waitForElementVisible(driver,(By.ID,zip_code_id))
        continueButton = self.waitForElementVisible(driver,((By.ID,continue_button_id)))
        actions.send_keys_to_element(firstName,"Muhsine")
        actions.send_keys_to_element(lastName,"Taşcı")
        actions.send_keys_to_element(zipCode,"123456")
        actions.click(continueButton)
        actions.perform()
        finishButton = self.waitForElementVisible(driver,(By.ID,finish_button_id))
        actions.click(finishButton).perform()
        sleep(2)
        baslik2 = self.waitForElementVisible(driver,(By.CLASS_NAME,baslik_class_name))
        testResult = baslik2.text == baslik2_text
        print(f"Ödeme işlemi başarılı: {testResult}")

    def test_dropDown(self):
        data = verileriKullanJson("C:/Users/hasine_tasci/OneDrive/Masaüstü/Selenium-Test-Tobeto/PyTest/data/veri.json")
        test_data = data.get("test_valid_login", {})  # JSON dosyasından verileri çeker
        username = test_data.get("username", "")  # Kullanıcı adını al
        password = test_data.get("password", "")  # Şifreyi al
        driver = self.precondition()
        userNameInput = self.waitForElementVisible(driver,(By.ID,username_id))
        passwordInput =self.waitForElementVisible(driver,(By.ID,password_id))
        actions = ActionChains(driver)
        actions.send_keys_to_element(userNameInput,username)
        actions.send_keys_to_element(passwordInput,password)
        loginButton = self.waitForElementVisible(driver,(By.ID,login_button_id))
        actions.click(loginButton)
        actions.perform() #depoladığım aksiyonları çalıştır
        #sıralam listesine tıklama
        dropDown = self.waitForElementVisible(driver,(By.CLASS_NAME,dropDown_class_name))
        actions.click(dropDown)
        actions.perform()
        sleep(3)
        #sıralama listesinden indexe göre seçim yapma
        secme = self.waitForElementVisible(driver,(By.CLASS_NAME,secme_class_name))
        select = Select(secme)
        select.select_by_index(3)
        sleep(2)
        # select.select_by_visible_text("Price (high to low)")
        # select.select_by_value("hilo")
        
    def test_menu(self):
        driver = self.precondition()
        data = verileriKullanJson("C:/Users/hasine_tasci/OneDrive/Masaüstü/Selenium-Test-Tobeto/PyTest/data/veri.json")
        test_data = data.get("test_valid_login", {})  # JSON dosyasından verileri çeker
        username = test_data.get("username", "")  # Kullanıcı adını al
        password = test_data.get("password", "")  # Şifreyi al
        userNameInput = self.waitForElementVisible(driver,(By.ID,username_id))
        passwordInput = self.waitForElementVisible(driver,(By.ID,password_id))
        actions = ActionChains(driver)
        actions.send_keys_to_element(userNameInput,username)
        actions.send_keys_to_element(passwordInput,password)
        loginButton = self.waitForElementVisible(driver,(By.ID,login_button_id))
        actions.click(loginButton)
        sleep(3)
        actions.perform()
        menuButton = self.waitForElementVisible(driver,(By.ID,menuButton_id))
        actions.click(menuButton).perform()
        aboutLink = self.waitForElementVisible(driver,(By.ID,aboutLink_id))
        actions.click(aboutLink)
        actions.perform()
        assert "saucelabs.com" in driver.current_url,"Yanlış sayfadasın."
        # print("Sauce Labs Sayfası açıldı.")
        sleep(2)